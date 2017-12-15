## Read western duel corpus workbook
# Read headers
# skip blank rows
# save and load data with pickle

from openpyxl import load_workbook
import pickle
from clockdeco import clock
from copy import deepcopy
from collections import defaultdict

EXCLUDE_SCENES = ['tg', None, 'none', 'None']

class Cell:
	def __init__(self, r, c):
		self._cell = [r, c]
		self.ledger = defaultdict()
	def __str__(self):
		return toStr(self._cell)
	def shiftRight(self):
		#self.setLast(deepcopy(self._cell))
		self._cell[0] = chr(ord(self._cell[0]) + 1)
	def shiftDown(self):
		#self.setLast(deepcopy(self._cell))
		self._cell[1] += 1
	def shiftDownRight(self):
		self.shiftRight()
		self.shiftDown()
	def remember(self, name, cell):
		self.ledger[name] = deepcopy(cell)
	def go(self, name):
		self._cell = self.ledger[name]

def toStr(tup):
	return str(tup[0]+str(tup[1]))

# an attempt to spit out new worksheet - needs improvement
def spit():
	newark = wb.create_sheet('system output', len(wb.worksheets)+1)
	scene_lib = load()
	cell = Cell('A', 1)
	for scene_name, scene in scene_lib.items():
		newark[str(cell)] = scene_name
		cell.shiftDownRight()
		next_shot = cell
		print(cell)
		for i, shot in enumerate(scene):
			cell = next_shot
			next_shot = deepcopy(cell)
			next_shot.shiftDown()
			cell.shiftDownRight()
			print(cell)
			for i, action in enumerate(shot.actions):
				newark[str(cell)] = str(action._type)
				cell.shiftRight()
				for arg in action:
					newark[str(cell)] = arg
					cell.shiftRight()


def clean(s):
	if not isinstance(s, str):
		return s
	return ''.join(s.split()).lower()

def toNumber(s):
	if s is None:
		return
	if not isinstance(s, str):
		print('not a string toNumber', s)
		return
	num_array = [int(i) for i in s if i.isdigit()]
	return int(''.join(str(i) for i in num_array))

def save_scenes(scene_lib, file_name):
	output = open(file_name, 'wb')
	pickle.dump(dict(scene_lib), output, protocol=pickle.HIGHEST_PROTOCOL)
	output.close()


class Header:
	def __init__(self, row_0):
		self._header = row_0

	def __len__(self):
		return len(self._header)

	def __getitem__(self, name):
		return self._header.index(name)

	@property
	def names(self):
		return self._header

	def fromTo(self, _from, _to):
		return (self[_from], self[_to]+1)

	def namesFromTo(self, _from, _to):
		return self._header[_from:_to]

	def __repr__(self):
		return self._header.__repr__()


class ActionType:
	def __init__(self, type_name, quant, old_name):
		self.type_name = type_name
		self.num_appearances = quant
		self._orig_name = old_name

	def updateQuant(self, new_quant_data):
		if self.type_name in new_quant_data.keys():
			self.num_appearances = int(new_quant_data[self.type_name])

	def asDict(self):
		return self.__dict__

	def __repr__(self):
		return str(self.type_name)

class Action:

	def __init__(self, **kwargs):
		self._type = kwargs['action']
		self._args = [kwargs['arguments']]
		self.starts = None
		self.finishes = kwargs['conclusionstatus']
		for key in kwargs:
			setattr(self, key, kwargs[key])

	def asDict(self):
		self._args = [arg.asDict() for arg in self._args if type(arg) is not str and arg is not None]
		# self._type = self._type.asDict()
		if type(self._type) is not str and self._type is not None:
			self._type = self._type.asDict()

		return self.__dict__

	def appendArg(self, arg):
		self._args.append(arg)

	def substituteEntities(self, ents):

		print('substituting entities {} in action {}\n'.format(ents, self._type))
		# print(self._args)
		try:
			new_args = []
			for arg in self:
				for e in ents:
					if e.name == arg:
						new_args.append(e)
						break
			self._args = new_args
		except:
			for e in ents:
				print(e)
			AttributeError('this should end it')
		return self._args

	def swap_arg(self, old_arg, new_arg):
		k = len(self._args)
		self._args = [arg if arg != old_arg else new_arg for arg in self]
		if len(self._args) != k:
			AssertionError('missing arg after swap: {}/{}'.format(old_arg, new_arg))

	def __len__(self):
		return len(self._args)

	def __getitem__(self, item):
		return self._args[item]

	def __repr__(self):
		args = str([arg for arg in self])
		return '{}'.format(self._type) + args


#store by scene name
import copy


class Shot:
	header = None
	def __init__(self, first_action, sentence, **kwargs):
		#self.__dict__.update(kwargs)
		for key in kwargs:
			setattr(self, key, kwargs[key])
		self.actions = [first_action]
		self.orig_sentence = sentence
		self.nlp_sentence = None
		self.assignSpatial()

	def asDict(self):
		self.actions = [action.asDict() for action in self.actions]
		return self.__dict__

	def assignSpatial(self):
		spatial_start = figure_split_machine(self.__dict__['figureobjs(shotstart)'])
		spatial_end = figure_split_machine(self.__dict__['figureobjs(shotend)'])
		spatials = spatial_start + spatial_end
		self.foreground = []
		self.background = []
		already_seen_ents = set()
		bkgrnd = {'background', 'back', 'behind', 'backgrond', 'bck', 'bbackground'}
		for ent in spatials:

			if len(ent) < 2:
				continue

			# if ent[0] == 'null':
			# 	continue
			#
			# try:
			if ent[1] in already_seen_ents:
				continue
			# except:
			# 	print('here')
			# 	continue
			ground = ent[-1].split('-')[-1]

			if ground in bkgrnd:
				self.background.append(ent[1:])
			else:
				self.foreground.append(ent[1:])

			already_seen_ents.add(ent[1])


	def update(self, row_values):
		#thus far, only update is to add argument to action
		arg = row_values[self.header['arguments']]
		self.actions[-1].appendArg(arg)

	def substituteEntities(self, ents):
		shot_ents = []
		# gather up the accepted substitutions in each action.
		for action in self.actions:
			accepted_ents = action.substituteEntities(ents)
			shot_ents.extend(accepted_ents)
		# should be sorta sorted by the order each entity is observed.
		self.entities = shot_ents

		ent_dict = dict()
		for ent in self.entities:
			ent_dict[ent.name.split('_')[0]] = ent

		for ent in self.foreground:
			try:
				if ent[0] in ent_dict.keys():
					ent[0] = ent_dict[ent[0]]
			except:
				print('here')
		for ent in self.background:
			if ent[0] in ent_dict.keys():
				ent[0] = ent_dict[ent[0]]



	def substituteConjunctionEntities(self, conj_ent_dict):

		for i, action in enumerate(self.actions):
			for arg in action:
				if arg in conj_ent_dict.keys():
					new_ents = conj_ent_dict[arg]
					for k, ent in enumerate(new_ents[1:]):
						new_action = copy.deepcopy(action)
						new_action.swap_arg(arg, ent)
						self.insert_action(i+k+1, new_action)
					# swap the ent in the first action by default.
					action.swap_arg(arg, new_ents[0])


	def insert_action(self, action, position):
		self.actions.insert(action, position)

	def __repr__(self):
		actions = [' '.join('\t' + str(i) + ': ' + str(action) for i, action in enumerate(self.actions))]
		return '\n' + ''.join(['{}'.format(action) for action in actions])

def figure_split_machine(_str):
	"""
	:param _str: a figureobjs(shotstart) string
	:return: a nested list where each primitive list is of the form ['type', 'ent_name', 'spatial position']
	"""

	ents = []
	if _str is None or _str == 'None' or _str == 'none':
		return ''
	str_list = _str.split(',')
	for i in str_list:
		if i == '':
			continue
		z = i.split('(')
		new_ent = []
		for k in z:
			if k == '':
				continue
			m = k.split(')')
			for q in m:
				if q == '':
					continue
				new_ent.append(q)
		ents.append(new_ent)
	return ents


class Scene:
	# name, ordered list of shots, entities

	def __init__(self, name):
		self._shots = []
		self.name = name
		self.entities = set()
		self.sentences = None

	def __len__(self):
		return len(self._shots)

	def __getitem__(self, item):
		return self._shots[item]

	def asDict(self):

		self._shots = [shot.asDict() for shot in self._shots]
		self.entities = [ent.asDict() for ent in self.entities if type(ent) is not str]
		return self.__dict__

	def append(self, item):
		self._shots.append(item)

	def addShot(self, shot):
		self._shots.append(shot)

	def substituteEntities(self, role_dict):
		print('substituting entities in scene {}'.format(self.name))

		conj_ents_dict = {e: role_dict[e] for e in self.entities if len(role_dict[e]) > 1}
		if conj_ents_dict:
			for shot in self:
				shot.substituteConjunctionEntities(conj_ents_dict)

		self.entities = [elm[0] for elm in role_dict.values() if len(elm) == 1]

		for shot in self:
			shot.substituteEntities(self.entities)

	def substituteActionTypes(self, action_dict, action_count):
		print('substituting action types in scenes {}'.format(self.name))
		for shot in self:
			for action in shot.actions:
				ac = action_count[action]

				if action._type in action_dict.keys():
					action_type = action_dict[action._type]
					action_type.num_appearances = ac
				else:
					if action._type is None or action._type == 'None' or action._type == 'none':
						action_type = ActionType(None, ac, action._type)
					else:
						action_type = ActionType(action._type, ac, action._type)

				action._type = action_type

	def __getattribute__(self, name):

		if name is 'shot' and not self._shots:
			print('no shot')
			return -1
		else:
			return object.__getattribute__(self, name)

	def __repr__(self):
		shots = ['\n' + str(i) + ':' + str(shot) for i, shot in enumerate(self)]
		return '\nSCENE:' + str(self.name) + '\n' + ''.join(shot for shot in shots)

# A class for storing scenes, can be forgotten and treated as a dictionary
class SceneLib:
	""" A mutable mapping / dictionary typed object
	# :warning ('fromKeys' not implemented)
	# :methods (.keys() and .values() and .items())
	"""

	def __init__(self, names):
		self._scenes = {name: Scene(str(name)) for name in names}

	def __len__(self):
		return len(self._scenes)

	def __getitem__(self, name):
		return self._scenes[name]

	def __setitem__(self, key, value):
		self._scenes[key] = value

	def __contains__(self, item):
		return item in self._scenes.keys()

	def __delitem__(self, key):
		del self._scenes[key]

	def __iter__(self):
		return iter(self._scenes.items())

	# def __str__(self):
	# 	return str(self._scenes)

	def keys(self):
		return self._scenes.keys()

	def values(self):
		return self._scenes.values()

	def items(self):
		return self._scenes.items()

	def asDict(self):
		return {key: value.asDict() for key, value in self._scenes.items()}

	def __repr__(self):
		members = ['\n' + str(value) for value in self.values()]
		return 'All Scenes: ' + '\n' + ''.join(['{}'.format(scene) for scene in members])

def load(d='scenelib.pkl'):
	return pickle.load(open(d, 'rb'))

def compileEntities(scene_lib):
	for scene in scene_lib.values():
		for shot in scene:
			for action in shot.actions:
				scene.entities.update({arg for arg in action._args if arg is not None})


def parse(ws):
	rows = list(ws.rows)
	header_rows = [clean(r.value) for r in rows[0]]
	header = Header(header_rows)
	Shot.header = header
	rows = [list(r) for r in rows[1:]]
	action_start, action_stop = header.fromTo('actionnumber', "conclusionstatus")

	# Scene Lib
	# global scene_lib
	scene_names = {clean(s.value) for s in list(ws.columns)[0][1:]}
	scene_lib = SceneLib(scene_names)

	print('starting parsing')

	last_shot_num = 0
	last_action_num = 0
	for row in rows:

		row_values = [clean(r.value) for r in row]
		scene_name = row_values[0]

		action_params = dict(zip(header.namesFromTo(action_start, action_stop), row_values[action_start:action_stop]))

		if len(row_values) != len(header):
			print("ALERT, |row_values| != |header|")

		elif toNumber(row_values[header['shotnumber']]) == last_shot_num:

			# same shot,
			last_shot = scene_lib[row_values[0]][-1]

			action_num = row_values[header['actionnumber']]
			if action_num != last_action_num:
				# new action
				last_shot.actions.append(Action(**action_params))
				last_action_num += 1

			else:
				# new action argument
				last_shot.update(row_values)
		else:
			# new shot, first action
			first_action = Action(**action_params)
			shot_desc = row[header['eventdescription']].value
			new_shot = Shot(first_action, shot_desc, **dict(zip(header.names, row_values)))
			scene_lib[scene_name].append(new_shot)

			if toNumber(row_values[header['shotnumber']]) == 1 and not last_shot_num == 0:
				#new scene, reset counter
				last_shot_num = 1
			else:
				last_shot_num += 1

			last_action_num = 1
	# print(scenes)
	print('compiling scene entities')
	compileEntities(scene_lib)
	# save_scenes(scene_lib)
	return scene_lib

def readCorpus(file_name='Western_duel_corpus_edited.xlsx'):
	print('loading workbook')
	wb = load_workbook(filename=file_name, data_only=True)
	return wb.worksheets[0]

if __name__ == '__main__':
	old_file_name = 'Western_duel_corpus_edited.xlsx'
	master_file_name = 'WESTERN_DUEL_CROPUS_MASTER_update.xlsx'
	rc = readCorpus(file_name=master_file_name)
	parse(rc)


